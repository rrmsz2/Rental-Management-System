from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any
from motor.motor_asyncio import AsyncIOMotorDatabase
import arabic_reshaper
from bidi.algorithm import get_display

# PDF libraries
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_RIGHT, TA_CENTER

# Excel libraries
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExportService:
    def __init__(self, db: AsyncIOMotorDatabase):
        self.db = db
    
    @staticmethod
    def reshape_arabic(text: str) -> str:
        """إعادة تشكيل النص العربي للعرض الصحيح"""
        if not text:
            return ""
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)
    
    # ========== PDF Export ==========
    
    async def export_invoice_pdf(self, invoice_id: str) -> BytesIO:
        """تصدير فاتورة واحدة بصيغة PDF"""
        # الحصول على الفاتورة
        invoice = await self.db.invoices.find_one({"id": invoice_id})
        if not invoice:
            raise ValueError("Invoice not found")
        
        # الحصول على العقد
        contract = await self.db.rental_contracts.find_one({"id": invoice["contract_id"]})
        
        # الحصول على العميل والمعدة
        customer = await self.db.customers.find_one({"id": contract["customer_id"]})
        equipment = await self.db.equipment.find_one({"id": contract["equipment_id"]})
        
        # إنشاء ملف PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        
        # العنوان
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        elements.append(Paragraph("INVOICE / فاتورة", title_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # معلومات الفاتورة
        info_data = [
            ["Invoice No:", invoice["invoice_no"], "رقم الفاتورة"],
            ["Issue Date:", invoice["issue_date"][:10], "تاريخ الإصدار"],
            ["Status:", "Paid" if invoice["paid"] else "Unpaid", "الحالة"],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 2*inch, 2*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3 * inch))
        
        # معلومات العميل
        customer_data = [
            ["Customer Information", "", "معلومات العميل"],
            ["Name:", customer["full_name"], "الاسم"],
            ["Phone:", customer["phone"], "الهاتف"],
            ["Address:", customer.get("address", "N/A"), "العنوان"],
        ]
        
        customer_table = Table(customer_data, colWidths=[2*inch, 2*inch, 2*inch])
        customer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(customer_table)
        elements.append(Spacer(1, 0.3 * inch))
        
        # تفاصيل الإيجار
        rental_data = [
            ["Rental Details", "", "تفاصيل الإيجار"],
            ["Equipment:", equipment["name"], "المعدة"],
            ["Contract No:", contract["contract_no"], "رقم العقد"],
            ["Start Date:", contract["start_date"][:10], "تاريخ البدء"],
            ["End Date:", contract.get("end_date", "N/A")[:10] if contract.get("end_date") else "N/A", "تاريخ الانتهاء"],
            ["Daily Rate:", f"{contract['daily_rate_snap']:.2f} OMR", "الأجرة اليومية"],
        ]
        
        rental_table = Table(rental_data, colWidths=[2*inch, 2*inch, 2*inch])
        rental_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(rental_table)
        elements.append(Spacer(1, 0.3 * inch))
        
        # الحسابات المالية
        amounts_data = [
            ["Financial Summary", "", "الملخص المالي"],
            ["Subtotal:", f"{invoice['subtotal']:.2f} OMR", "المبلغ الأساسي"],
            ["Tax ({:.1f}%):".format(invoice['tax_rate'] * 100), f"{invoice['tax_amount']:.2f} OMR", "الضريبة"],
            ["Discount:", f"{invoice['discount_amount']:.2f} OMR", "الخصم"],
            ["Total:", f"{invoice['total']:.2f} OMR", "الإجمالي"],
        ]
        
        amounts_table = Table(amounts_data, colWidths=[2*inch, 2*inch, 2*inch])
        amounts_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#dcfce7')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, -1), (-1, -1), 14),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(amounts_table)
        
        # Footer
        elements.append(Spacer(1, 0.5 * inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("Thank you for your business! | شكراً لتعاملكم معنا!", footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    # ========== Excel Export ==========
    
    async def export_rentals_excel(self, status: str = None) -> BytesIO:
        """تصدير قائمة العقود بصيغة Excel"""
        # الحصول على العقود
        query = {}
        if status:
            query["status"] = status
        
        rentals = await self.db.rental_contracts.find(query).to_list(1000)
        
        # إنشاء Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Rental Contracts"
        
        # تنسيق الرأس
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # الرأس
        headers = ["Contract No", "Customer", "Equipment", "Start Date", "End Date", "Daily Rate", "Status", "Created At"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # البيانات
        for row_num, rental in enumerate(rentals, 2):
            # الحصول على العميل والمعدة
            customer = await self.db.customers.find_one({"id": rental["customer_id"]})
            equipment = await self.db.equipment.find_one({"id": rental["equipment_id"]})
            
            ws.cell(row=row_num, column=1, value=rental["contract_no"])
            ws.cell(row=row_num, column=2, value=customer["full_name"] if customer else "N/A")
            ws.cell(row=row_num, column=3, value=equipment["name"] if equipment else "N/A")
            ws.cell(row=row_num, column=4, value=rental["start_date"][:10])
            ws.cell(row=row_num, column=5, value=rental.get("end_date", "N/A")[:10] if rental.get("end_date") else "N/A")
            ws.cell(row=row_num, column=6, value=rental["daily_rate_snap"])
            ws.cell(row=row_num, column=7, value=rental["status"])
            ws.cell(row=row_num, column=8, value=rental["created_at"][:10])
            
            # تلوين حسب الحالة
            status_cell = ws.cell(row=row_num, column=7)
            if rental["status"] == "active":
                status_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            elif rental["status"] == "closed":
                status_cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        
        # ضبط عرض الأعمدة
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # حفظ في buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    async def export_customers_excel(self) -> BytesIO:
        """تصدير قائمة العملاء بصيغة Excel"""
        customers = await self.db.customers.find({}).to_list(1000)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        
        # تنسيق الرأس
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # الرأس
        headers = ["Full Name", "Phone", "Email", "National ID", "Address", "WhatsApp Opt-in", "Created At"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # البيانات
        for row_num, customer in enumerate(customers, 2):
            ws.cell(row=row_num, column=1, value=customer["full_name"])
            ws.cell(row=row_num, column=2, value=customer["phone"])
            ws.cell(row=row_num, column=3, value=customer.get("email", ""))
            ws.cell(row=row_num, column=4, value=customer.get("national_id", ""))
            ws.cell(row=row_num, column=5, value=customer.get("address", ""))
            ws.cell(row=row_num, column=6, value="Yes" if customer.get("whatsapp_opt_in") else "No")
            ws.cell(row=row_num, column=7, value=customer["created_at"][:10])
        
        # ضبط عرض الأعمدة
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    async def export_equipment_excel(self) -> BytesIO:
        """تصدير قائمة المعدات بصيغة Excel"""
        equipment_list = await self.db.equipment.find({}).to_list(1000)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipment"
        
        # تنسيق الرأس
        header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # الرأس
        headers = ["Name", "Category", "Serial No", "Daily Rate", "Status", "Purchase Date", "Created At"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # البيانات
        for row_num, equipment in enumerate(equipment_list, 2):
            ws.cell(row=row_num, column=1, value=equipment["name"])
            ws.cell(row=row_num, column=2, value=equipment["category"])
            ws.cell(row=row_num, column=3, value=equipment.get("serial_no", ""))
            ws.cell(row=row_num, column=4, value=equipment["daily_rate"])
            ws.cell(row=row_num, column=5, value=equipment["status"])
            ws.cell(row=row_num, column=6, value=equipment.get("purchase_date", "")[:10] if equipment.get("purchase_date") else "")
            ws.cell(row=row_num, column=7, value=equipment["created_at"][:10])
            
            # تلوين حسب الحالة
            status_cell = ws.cell(row=row_num, column=5)
            if equipment["status"] == "available":
                status_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            elif equipment["status"] == "rented":
                status_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            elif equipment["status"] == "maintenance":
                status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        # ضبط عرض الأعمدة
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    async def export_revenue_report_pdf(self, start_date: str = None, end_date: str = None) -> BytesIO:
        """تصدير تقرير الإيرادات بصيغة PDF"""
        from services.reports_service import ReportsService
        
        reports_service = ReportsService(self.db)
        report = await reports_service.get_revenue_report(start_date, end_date)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        elements.append(Paragraph("Revenue Report / تقرير الإيرادات", title_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # الفترة
        period_text = f"Period: {start_date or 'All'} to {end_date or 'All'}"
        elements.append(Paragraph(period_text, styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))
        
        # الملخص
        summary = report["summary"]
        summary_data = [
            ["Metric", "Value", "المقياس"],
            ["Total Invoices", str(summary["total_invoices"]), "عدد الفواتير"],
            ["Total Revenue", f"{summary['total_revenue']:.2f} OMR", "إجمالي الإيرادات"],
            ["Total Tax", f"{summary['total_tax']:.2f} OMR", "إجمالي الضريبة"],
            ["Total Discount", f"{summary['total_discount']:.2f} OMR", "إجمالي الخصم"],
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(summary_table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
